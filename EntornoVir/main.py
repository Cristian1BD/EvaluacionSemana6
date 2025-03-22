import openpyxl

# Crear un nuevo archivo de Excel
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Estudiantes"

# Escribir el encabezado
sheet.cell(row=1, column=1).value = "Nombres cortos (<=4 letras)"

# Crear un diccionario de estudiantes
estudiantes = {}

# Solicitar nombres y notas al usuario
for i in range(3):
    nombre = input(f"Ingrese el nombre del estudiante {i+1}: ")
    nota = float(input(f"Ingrese la nota del estudiante {i+1}: "))
    estudiantes[nombre] = nota

# Escribir nombres cortos en la hoja de Excel
last_row = 2  # Comienza en la fila 2 después del encabezado
for nombre in estudiantes:
    if len(nombre) <= 4:  # Condición para nombres cortos
        sheet.cell(row=last_row, column=1).value = nombre
        last_row += 1

# Guardar los cambios en el archivo
workbook.save("ejercicio4.xlsx")
print("¡Ejercicio 4 guardado en ejercicio4.xlsx!")
